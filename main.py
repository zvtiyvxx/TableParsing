from openpyxl import load_workbook


filename = 'ScheduleFacult/uits24.xlsx'

wb = load_workbook(filename, data_only=True)
sheetnames = wb.sheetnames

def search_group(wb, group_name):
    for sheetname in sheetnames:
        sheet = wb[sheetname]
        row_number = 7
        row = sheet[row_number]
        for idx, cell in enumerate(row):

            if cell.value == group_name:
                print(f'Найдено в листе {sheet.title}, в ячейке {cell.coordinate}')
                if idx + 1 < len(row):
                    next_cell = row[idx + 1]
                    if next_cell.value is None:
                        print(f"В {group_name} есть 2 подгруппа")
                        return cell, sheet, next_cell
                    else:
                        print("Подгрупп не обнаружено")
                        return cell, sheet, None
                else:
                    print("Подгрупп не обнаружено")
                    return cell, sheet, None
    return None, None, None

def parsing(sheet, cell):
    days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    times = [
        "08:00 - 09:35", "09:45 - 11:20", "11:50 - 13:25",
        "13:35 - 15:10", "15:20 - 16:55", "17:05 - 18:40", "18:50 - 20:25"
    ]

    schedule = []

    def find_root_cell(sheet, cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, _, _ = range_.bounds
                return sheet.cell(row=min_row, column=min_col)
        return cell

    def is_merged_4_cells(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                min_col, min_row, max_col, max_row = range_.bounds
                if (max_col - min_col + 1) * (max_row - min_row + 1) == 4:
                    return True, min_col, min_row, max_col, max_row
        return False, None, None, None, None

    def is_merged(cell):
        for range_ in sheet.merged_cells.ranges:
            if cell.coordinate in range_:
                return True
        return False

    row = cell.row + 1
    col = cell.column

    for day in days:
        classes_count = 0
        while classes_count < 7:
            current_cell = sheet.cell(row=row, column=col)
            time_slot = times[classes_count]

            if is_merged_4_cells(current_cell)[0]:
                root_cell = find_root_cell(sheet, current_cell)
                cell_value = root_cell.value
                schedule.append({
                    "day": day,
                    "time": time_slot,
                    "числитель": cell_value,
                    "знаменатель": cell_value
                })
                row += 2
            else:
                chislitel_cell = sheet.cell(row=row, column=col)
                znamenatel_cell = sheet.cell(row=row + 1, column=col)

                chislitel_value = find_root_cell(sheet, chislitel_cell).value if is_merged(chislitel_cell) else chislitel_cell.value
                znamenatel_value = find_root_cell(sheet, znamenatel_cell).value if is_merged(znamenatel_cell) else znamenatel_cell.value

                schedule.append({
                    "day": day,
                    "time": time_slot,
                    "числитель": chislitel_value,
                    "знаменатель": znamenatel_value
                })
                row += 2

            classes_count += 1

        row = cell.row + 1 + ((days.index(day) + 1) * 14)  # Переход на следующий день

    return schedule


def parsing_2subgroups(sheet, cell, next_cell):
    subgroups_schedule = {
        'group_2': parsing(sheet, next_cell)
    }

    return subgroups_schedule

cell, sheet, next_cell = search_group(wb, 'У-232')
print(cell, next_cell)
subgroups_schedule = parsing_2subgroups(sheet, cell, next_cell)
print(subgroups_schedule)


def print_schedule(schedule_data):
    for group, schedule in schedule_data.items():
        print(f"\n=== Расписание для группы {group} ===\n")
        current_day = None

        for entry in schedule:
            day = entry["day"]
            time = entry["time"]

            # Проверяем значения на None и приводим их к строке
            chislitel = entry.get("числитель", "")
            chislitel = chislitel.strip() if isinstance(chislitel, str) else "—"

            znamenatel = entry.get("знаменатель", "")
            znamenatel = znamenatel.strip() if isinstance(znamenatel, str) else "—"

            # Печать заголовка дня только один раз
            if day != current_day:
                print(f"\n{day}:")
                current_day = day

            print(f"{time}:")
            print(f"  Числитель: {chislitel}")
            print(f"  Знаменатель: {znamenatel}")
            print("-" * 40)  # Разделитель между парами
print_schedule(subgroups_schedule)

